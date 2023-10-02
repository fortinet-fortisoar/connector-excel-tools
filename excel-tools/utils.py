import openpyxl
import os
from jsonschema import validate
from connectors.cyops_utilities.builtins import download_file_from_cyops, upload_file_to_cyops
from django.conf import settings
from connectors.core.connector import get_logger, ConnectorError

logger = get_logger('excel-tools')

def get_available_operations(operations, operation):
    '''returns the function object defined by operation str'''
    for func in filter(callable, operations.__dict__.values()):
        if operation in func.__qualname__:
            return func


def get_file_from_fortisoar(file_iri):
    try:
        cyops_file = download_file_from_cyops(file_iri)
        cyops_file_path = cyops_file.get('cyops_file_path')
        file_name = cyops_file.get('filename') if cyops_file.get('filename') else cyops_file.get('cyops_file_path')
        path_on_disk = os.path.join(settings.TMP_FILE_ROOT,cyops_file_path + '.xlsx') # new name
        cyops_file_path = os.path.join(settings.TMP_FILE_ROOT, cyops_file_path) # current name
        os.rename(cyops_file_path, path_on_disk)
        return path_on_disk, file_name
    
    except Exception as err:
        logger.exception('Could not download file: {} from FortiSOAR. {}'.format(file_iri, err))
        raise ConnectorError('Could not download file: {} from FortiSOAR. {}'.format(file_iri, err))


def load_workbook(params):
    try:
        file_iri = params.get('file_iri')
        sheet_name = params.get('sheet_name', None)
        file_path, file_name = get_file_from_fortisoar(file_iri)
        workbook = openpyxl.load_workbook(filename=file_path)
        if sheet_name:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                raise ConnectorError('File defined by :{0} has no sheet named: {1}'.format(file_iri, sheet_name))
        else:
            sheet = workbook.active
        
        return {"workbook": workbook, "sheet": sheet, "filePath": file_path, "fileName": file_name}
    except Exception as err:
        logger.exception('List Sheets Error: {0}'.format(err))
        raise ConnectorError('List Sheets Error: {0}'.format(err))


def cleanup(workbook):
    try:
        os.remove(workbook['filePath'])
    except OSError as err:
        logger.warning('Couldnd not delete file: {0}. Error: {1}'.format(workbook['filePath'], err))


def upload_file_to_fortisoar(workbook_dict, create_attachment=True):
    try:
        workbook = workbook_dict['workbook']
        file_name = workbook_dict['fileName']
        file_path = workbook_dict['filePath']
        workbook.save(filename=file_path)
        upload_response = upload_file_to_cyops(
            file_path=file_path,
            filename=file_name,
            name=file_name,
            create_attachment=create_attachment
            )
        file_iri = upload_response['file']['@id'] if create_attachment else upload_response['@id']

        cleanup(workbook_dict)
        return { "fileIRI": file_iri, "response": upload_response }

    except Exception as err:
        logger.warning('Couldnd upload file: {0} to FortiSOAR. Error: {1}'.format(file_path, err))
        raise ConnectorError('Couldnd upload file: {0} to FortiSOAR. Error: {1}'.format(file_path, err))
    
def validate_json_schema(_instance, _schema):
    try:
        validate(instance=_instance, schema=_schema)
        return _instance
    except Exception as err:
        logger.error("Entered data: [{0}] does not match the schema:{1}. Error: {2}".format(_instance,_schema,err))
        raise ConnectorError("Entered data: [{0}] does not match the schema:{1}. Error: {2}".format(_instance,_schema,err))
    

def parse_input_list(input_list):
    if isinstance(input_list, str) and ',' in input_list:
        input_list = input_list.replace(' ','')
        return input_list.split(',')
    elif isinstance(input_list, list):
        return input_list